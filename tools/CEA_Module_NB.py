"""
Purpose: The purpose is to check any input feature class for polygon overlap then output a new file or overwrite the
existing and clean up overlaps.

Date: October, 2013

Created by: Graham MacGregor Thompson Okanagan Region (FLNRO)

Arguments:

Version: created in arcgis 10.0 arcpy format

Classes
-----------
featureClass(The feature class to check, the name of the ouput feature class, the boundary area extent that will
be checked (if no boundary)
        -     this class will provide easy access to the properties of a featureClass.
            if no gp is provided the full path of the featureClass/layer/shapefile
            must be provided.

        -    Properties such as name, type, geomField, shapeType,
            and a list of field names are available

"""

# Run the regular stuff
import os, sys, arcpy, time
# sys.path.append(r'N:\FOR_RNI_RNI_Projects\WHPOR_Watershed_Analysis\1_WHPOR_Analyses\2023\6_Hominka\1_SpatialData\4_CEA_Watershed_Analysis\openpyxl')
import openpyxl 
from openpyxl.workbook import Workbook
from arcpy.sa.ParameterClasses import RemapRange
from arcpy.sa.Functions import ZonalStatisticsAsTable, Reclassify,\
    ExtractByMask, Slice
arcpy.env.overwriteOutput = True





# tempD=os.path.join(BaseFolder,r'1_SpatialData\4_CEA_Watershed_Analysis')

class gridClass:

    def __init__(self,BAfold): # , boundarySource):
        # self.boundarySource = boundarySource
        self.BAfold=BAfold
        
        tmpdrive = os.path.join(self.BAfold,r'1_SpatialData\4_CEA_Watershed_Analysis')
        tmpFGDB = r"temp_data"
        self.sExtractLoc = os.path.join(tmpdrive, tmpFGDB) + ".gdb"
        if not arcpy.Exists(self.sExtractLoc):
            arcpy.CreateFileGDB_management(tmpdrive, tmpFGDB)
            print ('Temporary file geodatabase created')
        arcpy.env.overwriteOutput = True

    def get_geom_envelope(self, inputFC):  # return Xmin Ymin Xmax Ymax
        print ('get geometry envelope')
        if arcpy.Exists(inputFC):
            # Local variables...
            fc = arcpy.Describe(inputFC)
            lyrextent = str(fc.extent)
            envelope = lyrextent.split(" ")
            xlist = [envelope[0], envelope[2]]
            ylist = [envelope[1], envelope[3]]
            xlist.sort()
            ylist.sort()
            print ('get geometry envelope complete')
            return str(xlist[0] + " " + ylist[0] + " " + xlist[1] + " " + ylist[1])  # may want to add 100m to extent
        
    def extractGrid(self, gridSource, outgrid, envelopeString=None, boundarySource=None): # , boundarySource, deleteDuplicate):
        print (boundarySource)
        print (envelopeString)
        if boundarySource != None:
            print ('Envelope string not available using layer to find Boundary source')
            self.get_geom_envelope(boundarySource)
            arcpy.Clip_management(gridSource, envelopeString, outgrid)
        if envelopeString != None:
            print ('Clipping Grid')
            arcpy.Clip_management(gridSource, envelopeString, outgrid)
            print ('Creating output Grid ' + outgrid)
            
    def categorizeGrid(self, inGrid, reclassrange, outputName): # , boundarySource, deleteDuplicate):
        print ('reclassifying slope')
        arcpy.CheckOutExtension("3D")
        arcpy.Reclassify_3d(inGrid, "VALUE", reclassrange, outputName, "DATA")
        arcpy.CheckInExtension("3D")
        
    def Grid_to_Poly(self, inGrid, gridval, outputName, cliptoAOI=None): # , boundarySource, deleteDuplicate):
        print ('converting grid to polygon')
        if cliptoAOI == None:
            arcpy.RasterToPolygon_conversion(inGrid, outputName, "NO_SIMPLIFY", gridval)
        if cliptoAOI != None:
            if arcpy.Exists(os.path.join(self.sExtractLoc, 'tmpgridpoly')):
                tmpFileloc = os.path.join(self.sExtractLoc, 'tmpgridpoly')+'1'
            else:
                tmpFileloc = os.path.join(self.sExtractLoc, 'tmpgridpoly')
            arcpy.RasterToPolygon_conversion(inGrid, tmpFileloc, "NO_SIMPLIFY", gridval)
            arcpy.Clip_analysis(tmpFileloc, cliptoAOI, outputName)


class extractData:
    
    def __init__(self): # , boundarySource):
        print ('self')
        
    def polygoncounter(self, selectedFeat,outputFeat):
        resultselfeat = arcpy.GetCount_management(selectedFeat)
        countselfeat = int(resultselfeat.GetOutput(0))
        resultoutfeat = arcpy.GetCount_management(outputFeat)
        countoutfeat = int(resultoutfeat.GetOutput(0))
        if countselfeat == countoutfeat:
            print ("input and output features are equal no need to worry")
        if countselfeat != countoutfeat:
            print ("Warning input features does not equal output selected features")
            sys.exit
    
    def extract_by_Location (self, inFeatureClass, AOI, outFeatureClass, SQL=None):
        if arcpy.Exists(inFeatureClass):
            arcpy.MakeFeatureLayer_management(inFeatureClass, "featLyr")
        # Extract by area - Clips inFeatureCLass to AOI and writes output to outFeatureClass
        selFeat = arcpy.SelectLayerByLocation_management("featLyr",'INTERSECT', AOI)
        # get count of number selected
        arcpy.CopyFeatures_management(selFeat, outFeatureClass)
        arcpy.Delete_management("featLyr")
        # checks that selected features output to a feature are equal.
        # self.polygoncounter("featLyr",outFeatureClass)
        
    def extract_by_Distance (self, inFeatureClass, AOI, outFeatureClass, Distance, SQL=None):
        if arcpy.Exists(inFeatureClass):
            arcpy.MakeFeatureLayer_management(inFeatureClass, "featLyr")
        # Extract by area - Clips inFeatureCLass to AOI and writes output to outFeatureClass
        selFeat = arcpy.SelectLayerByLocation_management("featLyr",'WITHIN_A_DISTANCE', AOI, Distance)
        # get count of number selected
        arcpy.CopyFeatures_management(selFeat, outFeatureClass)
        arcpy.Delete_management("featLyr")
        # checks that selected features output to a feature are equal.
        # self.polygoncounter("featLyr",outFeatureClass)
    def extract_by_Clip (self, inFeatureClass, AOI, outFeatureClass, SQL=None):
        if arcpy.Exists(inFeatureClass):
            arcpy.MakeFeatureLayer_management(inFeatureClass, "featclpLyr")
        # Extract by area - Clips inFeatureCLass to AOI and writes output to outFeatureClass
        selFeat = arcpy.SelectLayerByLocation_management("featLyr",'INTERSECT', AOI)
        # get count of number selected
        tmpFeat = arcpy.CopyFeatures_management(selFeat, os.path.join(self.sExtractLoc,outFeatureClass))
        arcpy.Clip_analysis(tmpFeat, AOI, outFeatureClass)
        # checks that selected features output to a feature are equal.
        # self.polygoncounter("featLyr",outFeatureClass)
        
    def return_list_items_in_field (self, inFeatureClass, inField):
        theList = []
        rows = arcpy.SearchCursor(inFeatureClass)
        for row in rows:
            if row.getValue(inField) not in theList:
                theList.append(row.getValue(inField))
        del row, rows
        return theList
    
    def extract_by_Mapsheet (self, inMapsheet, AOI, outFeatureClass, inMapList, SQL=None):
        print ('Extract by mapsheet')
        
    def buffer_Featureclass (self, inMapsheet, AOI, outFeatureClass, SQL=None):
        print ('Buffer feature class')


class analysis_utils:
    
    def __init__(self):
        print ('self')
        
    def append_data (self, indataLocation, inAppend_dataname, inbaseName, inList, dataType):
        appendList = []
        outAppenddata = os.path.join(indataLocation,inAppend_dataname)
        for name in inList:
                Datatemplate = os.path.join(indataLocation,inbaseName + name)
                #will create append file based on first template
                if not arcpy.Exists(outAppenddata):
                    print ('Creating Feature class based on first feature ' + Datatemplate)
                    print ('Append feature will be called ' + inAppend_dataname)
                    arcpy.CreateFeatureclass_management(indataLocation, inAppend_dataname, dataType, Datatemplate)
                if Datatemplate not in appendList:
                    appendList.append(Datatemplate)
        arcpy.Append_management(appendList,outAppenddata,'TEST')


class featureclass_utils:
    inTable_fieldList = []
    
    def __init__(self):
        print ('self')
        
    def GetGeometryField(self, inFC):
        desc = arcpy.Describe(inFC)
        ftype = desc.ShapeType
        if ftype == 'Polygon':
            areaItem = desc.areaFieldName
            lengthItem = desc.lengthFieldName
        if ftype == 'Polyline':
            areaItem = desc.lengthFieldName
            lengthItem = areaItem
        return areaItem, lengthItem

    def join_table(self, inTable, injoinField, joinTable, joinField, fieldList=None):
        if fieldList == None:
            print ('Joining ' + joinTable + ' to table ' + inTable)
            arcpy.JoinField_management (inTable, injoinField, joinTable, joinField, fieldList)
        if fieldList != None:
            print ('Joining ' + joinTable + ' to table ' + inTable + " with the fields ")
            arcpy.JoinField_management (inTable, injoinField, joinTable, joinField, fieldList)

    def delete_fields(self, inTable, inKeepList):
            fields = arcpy.ListFields(inTable)
            inTable_fieldList = []
            for field in fields:
                inTable_fieldList.append(field.name)
            del field, fields
            # del FTEN fields and keep which ones we want
            for field in inTable_fieldList:
                if field != 'OBJECTID':
                    if field not in inKeepList:
                        arcpy.DeleteField_management(inTable, field)
                    
    def return_field_list(self, listTable):
            fields = arcpy.ListFields(listTable)
            for field in fields:
                self.inTable_fieldList.append(field.name)
            return self.inTable_fieldList
        
                    
class table_utils:
    
    def __init__(self):
        print ('self')
        
    def Populate_table_withdictionary(self, inTable, fieldtopopulate, keyfield, inDictionary):
        cursor = arcpy.UpdateCursor(inTable)
        for row in cursor:
            keyvalue = row.getValue(keyfield)
            row.setValue(fieldtopopulate, inDictionary[keyvalue])
            cursor.updateRow(row)
            
    def zero_null_values(self, inDataset):
        desc = arcpy.Describe(inDataset)
        dataType = desc.DatasetType
        # field_names = [f.name for f in arcpy.ListFields(inDataset)]
        if dataType == 'FeatureClass':
            arcpy.MakeFeatureLayer_management(inDataset,'inLayer')
        if dataType == 'Table':
            arcpy.MakeTableView_management(inDataset,'inLayer')
        for field in desc.fields:
            name = field.name
            type = field.type
            arcpy.SelectLayerByAttribute_management('inLayer','NEW_SELECTION','"' + name + '"' + " is NULL")  #or blank?
            count = int(arcpy.GetCount_management('inLayer').getOutput(0))
            print ('Nulls selected for '+name+':  ' + str(count))
            if count > 0:
                if type != 'Text':
                    # Python 64
                    arcpy.CalculateField_management('inLayer', name , "0", "PYTHON_9.3", "")
                if type == 'Text':
                    arcpy.CalculateField_management('inLayer', name , "\"na\"")
            
class FGDB_utils:
    
    # def __init__(self):
    #     print ('self')
    def __init__(self,BAfold): # , boundarySource):
        self.BAfold=BAfold
        
    def make_FGDB(self, inOutputLoc, setOutputFGDB, inFeaturedataset = None):
        outFGDB = os.path.join(inOutputLoc,setOutputFGDB)+".gdb"
        outlocation = outFGDB
        if not arcpy.Exists(outFGDB):
            print ('Creating FGDB '+ outFGDB)
            arcpy.CreateFileGDB_management(inOutputLoc, setOutputFGDB)
        else:
            print (outFGDB + ' already exists not recreating')
        if inFeaturedataset != None:
            outFDS = os.path.join(outFGDB,inFeaturedataset)
            if not arcpy.Exists(outFDS):
                arcpy.env.XYTolerance = "0.01 Meters"
                arcpy.env.ZTolerance = "0.01 Meters"
                prj = os.path.join(self.BAfold, r"1_SpatialData\4_CEA_Watershed_Analysis\PCS_Albers.prj")
                arcpy.CreateFeatureDataset_management(outFGDB, inFeaturedataset, prj)
            outlocation = outFDS
        return outlocation
        
class watershedData:
    minDict = {}
    maxDict = {}
    
    def __init__(self):
        print ('self')

    def H_watershed_gen(self, wfcName, wfcNamefield, hType, OutHPoly_FD, DEMsource):
        try:
            from arcpy import sa
            arcpy.CheckOutExtension("Spatial")
        except:
            arcpy.AddError("Spatial Extension could not be checked out")
            os.sys.exit(0)

        # check to see if wfcNamefield exists and DEM source
        if not arcpy.Exists(wfcName):
            print(" This Feature does not exist exiting " + wfcName)
            sys.exit
        if not arcpy.Exists(DEMsource):
            print(" This Feature does not exist exiting " + DEMsource)
            sys.exit
        
        # list fields in watershed feature
        resultantfields = arcpy.ListFields(wfcName)
        tempfieldList = []
        for field in resultantfields:
            tempfieldList.append(field.name)
            
        # check to see if watershed name/ID exists if not exit
        if wfcNamefield not in tempfieldList:
            print ('the watershed name field does not exist in the watershed feature')
            sys.exit 

        # make list of watersheds
        print ('Creating list of watershed units')
        wsList = []
        rows = arcpy.SearchCursor(wfcName)
        for row in rows:
            if row.getValue(wfcNamefield) not in wsList:
                wsList.append(row.getValue(wfcNamefield))
        del row, rows
    
        try:
            print (hType)
            if hType == 'H40':
                print ('H40 was chosen')
                sliceZones = 5
                remap = RemapRange([[0, 3, 1], [4, 5, 2]])
                classNum = 2
                print (remap)
                lowZone = 'Lower60'
                uppZone = 'Upper40'
                print( uppZone)
            if hType == 'H50':
                sliceZones = 2
                remap = RemapRange([[0, 1, 1], [2, 2, 2]])
                classNum = 2
                lowZone = 'Lower50'
                uppZone = 'Upper50'
            if hType == 'H55':
                sliceZones = 20
                remap = RemapRange([[0, 9, 1], [10, 20, 2]])
                classNum = 2
                lowZone = 'Lower45'
                uppZone = 'Upper55'
            if hType == 'H60':
                print ('H60 was chosen')
                sliceZones = 5
                remap = RemapRange([[0, 2, 1], [3, 5, 2]])
                classNum = 2
                print (remap)
                lowZone = 'Lower40'
                uppZone = 'Upper60'
                print (uppZone)
            if hType == 'H70':
                sliceZones = 10
                remap = RemapRange([[0, 3, 1], [4, 10, 2]])
                classNum = 2
                print (remap)
                lowZone = 'Lower30'
                uppZone = 'Upper70'
                print( uppZone)
            # Will break into 3 classes: 0-30 (H70), 30-70, 70-100 (H40)
            if hType == 'H70_40':   
                sliceZones = 10
                remap = RemapRange([[0, 3, 1], [4, 6, 2], [7, 10, 3]])
                classNum = 3
                print (remap)
                lowZone = 'Lower30'
                midZone = 'Mid 30-70'
                uppZone = 'Upper40'
                print (uppZone)
        except:
                arcpy.AddError("Please enter appropriate watershed break type (H70,H60,H50,H55, H40, H70_40)")
                os.sys.exit(0)
        
        print ('Checking existence of individual Watershed H polys...')
        for watershed in wsList:
            # the name of the Hpoly layer
            Outws_FC = OutHPoly_FD+'\\'+hType+'_'+str(int(watershed))
            # print '\n'+ Outws_FC
            # if arcpy.Exists(Outws_FC):
            #    arcpy.Delete_management(Outws_FC)
            if not arcpy.Exists(Outws_FC):
                print ('\n Creating:  ' + Outws_FC)
                arcpy.env.cellSize = 100
                arcpy.AddMessage("Finding Watershed division elevation...")
                # query = 'Reporting_ = '+ str(uID)
                # wfcNamefield is 
                # query = '"' + wfcNamefield  + '"' + " = " + str(int(watershed))
                query = wfcNamefield + " = " + str(int(watershed))  # Assumes Unit ID is numeric
                print (query)
                arcpy.MakeFeatureLayer_management(wfcName, 'TempLyr1', query)
                count = int(arcpy.GetCount_management('TempLyr1').getOutput(0))   # the count should only be 1
                print ('Features Selected for DEM Mask:  ', count)
                arcpy.Clip_management(DEMsource, "", "DEM_Mask", "TempLyr1", "", "ClippingGeometry")
                # tDEM = ExtractByMask(DEMsource, 'TempLyr1')  # This takes a long time! Try setting mask environment
                print ('Extracted DEM for watershed:  ', str(int(watershed)))
                
                # Create dictionaries for min max of elevation

                min_val = arcpy.GetRasterProperties_management("DEM_Mask", "MINIMUM")
                self.minDict[watershed] = min_val
                max_val = arcpy.GetRasterProperties_management("DEM_Mask", "MAXIMUM")
                self.maxDict[watershed] = max_val
                print ('Min and Max Values: ', min_val, max_val)
                
                # arcpy.env.mask = 'TempLyr1' #Mask environment does not seem to work with Slice
                print ('Slice Zones: ' + str(sliceZones))
                tDemSlice = Slice("DEM_Mask", sliceZones, "EQUAL_AREA")
                # tDemSlice = Slice(DEMsource, sliceZones, "EQUAL_AREA")
                tDemSlice.save("temp_tDEMslc")

                # Reclass slices based on the Remap criteria
                arcpy.AddMessage("Reclassifying...")
                tClass = Reclassify('temp_tDEMslc', "Value", remap, "DATA")
                # Calc stats to get min/ max elevation in each zone etc
                ZonalStatisticsAsTable(tClass, "Value", "DEM_Mask", "temp_tBasinHStats")
                arcpy.Delete_management("DEM_Mask")
                arcpy.AddMessage("Converting to polys and adding Attributes...")
                # convert reclassified raster to polygons for the uppper watershed%
                arcpy.RasterToPolygon_conversion(tClass, "temp_tHDIV", "SIMPLIFY", "VALUE")
                # create multipart poly by dissolving

                # library location
                arcpy.Dissolve_management('temp_tHDIV', Outws_FC, 'gridcode', "", "MULTI_PART", "DISSOLVE_LINES")
                # join the stats table
                arcpy.JoinField_management(Outws_FC, 'gridcode', 'temp_tBasinHStats', 'VALUE', "MIN;MAX;RANGE")

                arcpy.AddField_management(Outws_FC, "Unit_ID", "DOUBLE", "", "", "15")  # Unique Watershed Unit Number
                arcpy.AddField_management(Outws_FC, "H_Line", "TEXT", "", "", "15")

                arcpy.CalculateField_management(Outws_FC, "Unit_ID", watershed, "PYTHON_9.3")
                arcpy.CalculateField_management(Outws_FC, "H_Line", '"' + hType+'"', "PYTHON_9.3")

                # Calc Area_ha
                print ('Calculating Area_ha...')
                if not arcpy.ListFields(Outws_FC, 'H_Line_Area_ha'):
                    arcpy.AddField_management(Outws_FC, 'H_Line_Area_ha', "DOUBLE", "", "2")
                arcpy.CalculateField_management(Outws_FC, "H_Line_Area_ha", "!shape.area@HECTARES!", "PYTHON_9.3")

                # Add Zone Field
                arcpy.AddField_management(Outws_FC, "Zone", "TEXT", "", "", "15")
                arcpy.MakeFeatureLayer_management(Outws_FC, "tmpLyr")
                arcpy.SelectLayerByAttribute_management('tmpLyr', "NEW_SELECTION", "\"gridcode\" = 1")
                arcpy.CalculateField_management('tmpLyr', "Zone", '"' + lowZone + '"', "PYTHON_9.3")

                if classNum == 2:
                    arcpy.SelectLayerByAttribute_management('tmpLyr', "NEW_SELECTION", "\"gridcode\" = 2")
                    arcpy.CalculateField_management('tmpLyr', "Zone", '"' + uppZone + '"', "PYTHON_9.3")

                if classNum == 3:
                    arcpy.SelectLayerByAttribute_management('tmpLyr', "NEW_SELECTION", "\"gridcode\" = 2")
                    arcpy.CalculateField_management('tmpLyr', "Zone", '"' + midZone + '"', "PYTHON_9.3")
                    arcpy.SelectLayerByAttribute_management('tmpLyr', "NEW_SELECTION", "\"gridcode\" = 3")
                    arcpy.CalculateField_management('tmpLyr', "Zone", '"' + uppZone + '"', "PYTHON_9.3")

                arcpy.SelectLayerByAttribute_management('tmpLyr', "CLEAR_SELECTION")
                
            else:
                print (Outws_FC+' Already Exists\n')
                # need to add min max to dictionary
                arcpy.MakeFeatureLayer_management(Outws_FC, "tmpLyr")
                # get min max elevation from current H_poly
                maxVal = 0
                minVal = 10000
                with arcpy.da.UpdateCursor(Outws_FC, ("MAX", "MIN", "Unit_ID")) as cursor:
                    for row in cursor:
                        maxRow = row[0] 
                        minRow = row[1]
                        if maxRow > maxVal:
                            maxVal = maxRow
                        if minRow < minVal:
                            minVal = minRow
                        row[2] = int(watershed)
                        cursor.updateRow(row)
                        
                #take the min max values and add them to the dictionary
                self.minDict[watershed] = minVal
                self.maxDict[watershed] = maxVal
        # this will take
        append_list = []
        outAppendpoly = "Append_Hpoly"+hType
        if not arcpy.Exists(OutHPoly_FD+'\\'+"Append_Hpoly"+hType):
            for watershed in wsList:
                append_list.append(str(int(watershed)))
            objappend = analysis_utils()
            objappend.append_data(OutHPoly_FD, outAppendpoly, hType+'_', append_list, "POLYGON")
        else:
            print ('The following append coverage already exists not recreating' + "Append_Hpoly"+hType)
                    
        # Cleanup
        print ('...file cleanup...')
        for Dfile in ('TempLyr1', 'tmpLyr', 'temp_tDEMslc', 'temp_tBasinHStats', 'temp_tHDIV'):
            try:
                arcpy.Delete_management(Dfile)
            except:
                arcpy.GetMessages()
        
        # Reset mask to entire area
        arcpy.env.mask = wfcName
        
        print ('Done H Polys for '+wfcName)
        # so Hpoly is not needed to be run from scratch make it into table form
        
        return self.minDict, self.maxDict

class excel_utils:

    def __init__(self):
        print ('self')
        
    def exportTable (self, inTableList, outFolder, outFilename):
        wb = Workbook() #Starts creating a workbook in memory
        dest_filename = os.path.join(outFolder,outFilename)+".xlsx" #The source location where you want the excel sheet to be
        sheet_num = 1
        for varType in inTableList: #The list layer helps do may files
            ws = wb.create_sheet(str(varType))#Makes a new sheet in the excel with the name ,sheet_num
            sheet_num = sheet_num + 1 #counts up one sheet for the next table to go in (for next loop)
            #ws.title = varType + "statistics"
            fieldnameslist = [f.name for  f in arcpy.ListFields(varType)] #Lists fields in in feature or table
            fieldnumvar = 0 #The first column is 0
            for i in fieldnameslist: #loops through field name list
                ws.cell(row = 0, column = fieldnumvar).value = i #goes just on first row and places field name
                fieldnumvar = fieldnumvar + 1 #move over one column
                colnum = -1 # sets first column back from 0
            for i in fieldnameslist: #starts loops through field name list
                colnum = colnum + 1 #adds 1 to column
                rownum = 1 #the starting row (0 is the column names)
                rows = arcpy.UpdateCursor(varType) #start cursor on the table
                for row in rows: #loops through each row for the set field name and adds the data to sheet
                    cellvalue = row.getValue(i)#gets value for the field name
                    ws.cell(row = rownum, column = colnum).value = cellvalue#calculates the cell to whats in the table
                    rownum = rownum + 1 #adds 1 to row for next loop and adding data
        wb.save(filename = dest_filename)#saves and makes destination file previously stated.

