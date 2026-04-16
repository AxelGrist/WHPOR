'''
=============================================================================================================
|   Final output script for the WHPOR project consits of six functions,                                    |
|   Deliverables are listed at the end of the script                                                        |
|                                                                                                           |
|   Fucntions                                                                                               |
|   1. move_CEA -   moves product from watershed analysis script to final locations                         |
|   2. xing_eca -   Calculates ECA and number of road/stream crossings                                      |
|   3. build_xlsx - Creates final compiled spreadsheet                                                      |
|   4. rejoin -     Joins all final calculations from spreadsheet back to spatial data                      |
|   5. maps -       Updates map template, adjusts according to watershed size and exports pdf               |
|   6. copDevs -    Creates folder in deliverable area and copies the speadsheet and map
| Created by C.Folkers                                                                                      |
| Date updated : 2023/07/20                                                                                 |
=============================================================================================================
'''
import arcpy
import os 
from glob import glob
import shutil 
import errno
import pandas as pd
import openpyxl as opxl
import datetime
import math 
import sys
import win32com.client as win32

class Results:
    def __init__(self, wtrshdname, Bfold):
        self.wtrshdname=wtrshdname
        self.Bfold=Bfold

        #user Variables
        WatershedName=self.wtrshdname
        BaseFolder=self.Bfold
        #static variables
        watershedname=WatershedName.replace(' ','_')
        today=datetime.datetime.today().strftime(r'%Y%m%d')
        year=str(datetime.datetime.today().year)
        unq_fol=BaseFolder.split("\\")[-1]
        rsltGdb=os.path.join(BaseFolder,r'1_SpatialData\4_CEA_Watershed_Analysis\Ouput\Compiled_Watershed_Hazard_Summaries_rw.gdb')
        outrslt=os.path.join(BaseFolder,r'1_SpatialData\3_ResultantData\Compiled_Watershed_Hazard_Summaries_rw.gdb')
        #r'N:\FOR_RNI_RNI_Projects\WHPOR_Watershed_Analysis\1_WHPOR_Analyses\2023\6_Hominka\1_SpatialData\3_ResultantData\Named\Compiled_Watershed_Hazard_Summaries_rw.gdb'
        inputgdb=os.path.join(BaseFolder,r'1_SpatialData\1_InputData',(watershedname+'_Input_Data.gdb'))
        #r'N:\FOR_RNI_RNI_Projects\WHPOR_Watershed_Analysis\1_WHPOR_Analyses\2023\6_Hominka\1_SpatialData\1_InputData\Hominka_River_Input_Data.gdb'
        rprtFolder=os.path.join(BaseFolder,r'2_Reports')
        # r'N:\FOR_RNI_RNI_Projects\WHPOR_Watershed_Analysis\1_WHPOR_Analyses\2023\6_Hominka\2_Reports'
        xlsTemplate=r'\\spatialfiles.bcgov\Work\for\RNI\RNI\Projects\WHPOR_Watershed_Analysis\working\source_data\Compiled_Watershed_Hazard_Summaries_Master8.xlsx'
        tempname=(watershedname+r'_Compiled_Watershed_Hazard_Summaries_'+today+r'.xlsx')
        report_out=os.path.join(rprtFolder,tempname)
        report_out2=os.path.join(rprtFolder,(r'Compiled_Watershed_Hazard_Summaries_'+WatershedName+r'_JOINS.xlsx'))
        aprxname=os.path.join(BaseFolder,r'1_SpatialData\1_InputData',(watershedname+'.aprx'))
        # aprxtemp=r'N:\FOR_RNI_RNI_Projects\WHPOR_Watershed_Analysis\!WHPOR_Stage\2_WHPOR_Resources\1_Templates_Utilities\2_APRX_Template\WHPOR_APRX_Template_20230713'
        aprxtemp=r'\\spatialfiles.bcgov\Work\for\RNI\RNI\Projects\WHPOR_Watershed_Analysis\working\source_data\WHPOR_APRX_Template_20230713\WHPOR_APRX_Template_20230713.aprx'
        mapname=(WatershedName+r' WHPOR Results Map '+today+'.pdf')
        mapout=os.path.join(BaseFolder,r'3_Maps', mapname)
        arcpy.env.overwriteOutput = True
        clientdir=os.path.join(r'\\spatialfiles.bcgov\Work\for\RNI\RNI\Projects\WHPOR_Watershed_Analysis',year,unq_fol)

        #function to spatial join xings and ECA to watersheds
        def move_CEA(rslt):
            arcpy.management.Copy(rslt,outrslt)
            # try:
            #     shutil.copytree(rslt,outrslt)
            # except OSError as exc: # python >2.5
            #     if exc.errno == errno.ENOTDIR:
            #         shutil.copy(rslt,outrslt)
            #     else: raise
            # arcpy.management.Copy(rslt,outrslt)
            # arcpy.management.Delete(rslt)
            print('Move That Bus!!!!!')
            arcpy.env.workspace=outrslt
            wtrshds=arcpy.ListFeatureClasses('*Compiled_Watershed_Features')
            print(wtrshds)
            

        def xing_eca (rslt,inpt):
        # xing_eca (outrslt,inputgdb)
            
            arcpy.env.workspace = rslt
            arcpy.env.overwriteOutput = True
            wtrshds=arcpy.ListFeatureClasses('Compiled_Watershed_Features*') 
            # wtrshds=arcpy.ListFeatureClasses('*Watersheds_in_AOI')
            tbls=arcpy.ListTables('Compiled_Watershed_Stats*')
            print(wtrshds)
            print(tbls)
            arcpy.env.workspace = inpt
            arcpy.env.overwriteOutput = True
            eca=arcpy.ListFeatureClasses('ECA')
            eca=eca[0]
            print(eca)
            #change to XINGS after 
            xing=arcpy.ListFeatureClasses('XINGS')
            xing=xing[0]
            print(xing)
            ecaha='ECA_ha'
            exprha='(!Shape_Area! /10000) * (!ECA_Factor!/100)'
            rev='RevRepUni'
            ecasc='ECA_Score'
            exprsc='(!SUM_ECA_ha! / !RU_Area_ha!)*100'
            eca_f_n_exp="fix_null(!ECA_Score!)"
            eca_fix_nulls='''def fix_null (eca_s):
            if eca_s is None:
                return 0
            else:
                return eca_s'''
            ecaR='ECA_Rank'
            expression='ECA_calc(!ECA_Score!)'
            codeblk='''def ECA_calc(value):
                if value >= 50:
                    return "VH"
                elif value >= 30:
                    return "H"
                elif value >= 20:
                    return "M"
                else:
                    return "L"
                    '''
            xcol='No_Crossings'
            for wtr,tbl in zip(wtrshds,tbls):
                arcpy.env.workspace = inpt
                print('======== Now Processing ========')
                print(wtr, ' and ', tbl)
                print('======== Now Processing ========')
                print(xing)
                inpath=os.path.join(rslt,wtr)
                outname1=wtr+'_XINGS'
                outname2=wtr+'_ECA_Intersect'
                outtable=wtr+'ECA_Summary'
                outname3=wtr+'not_final'
                outname3paht=os.path.join(rslt,outname3)
                outxls=os.path.join(rprtFolder,tbl+'.xlsx')
                if arcpy.Exists(outname1):
                    print (outname1, ' Already Exists')
                else:
                    arcpy.analysis.SpatialJoin(target_features=inpath, join_features= xing, out_feature_class=outname1, join_operation='JOIN_ONE_TO_ONE', join_type='KEEP_ALL',
                                        match_option='INTERSECT')
                print('spatial join xings')


                # intersect AOI with ECA so ECA has AOI Values
                if arcpy.Exists(outname2):
                    print(outname2, ' Already Exists')
                else:
                    arcpy.analysis.Intersect([outname1,eca],outname2)
                    print('Intersect ECA')
                    arcpy.management.AddField(outname2,ecaha,'DOUBLE')
                    arcpy.management.CalculateField(outname2,ecaha,exprha)
                    print('calculate ECA ha')
                if arcpy.Exists(outtable):
                    print(outtable,' already exists')
                else:
                    arcpy.analysis.Statistics(outname2,outtable,[[ecaha, "SUM"]],rev)
                    print('ECA Summary Table')
                if arcpy.Exists(outname3paht):
                    print(outname3paht, ' already exists')
                else:
                    arcpy.management.ValidateJoin(inpath,rev,outtable,rev)
                    arcpy.management.JoinField(inpath,rev,outtable,rev,['SUM_ECA_ha'])
                    arcpy.management.ValidateJoin(inpath,rev,outname1,rev)
                    arcpy.management.JoinField(inpath,rev,outname1,rev,['Join_Count'])
                    print('join ECA ha and xings Join Count to watershed resultant')
                    arcpy.conversion.ExportFeatures(inpath,outname3paht)
                    print('data exported to resultant gdb')
                    arcpy.management.AddField(outname3paht,ecasc,'DOUBLE')
                    arcpy.management.CalculateField(outname3paht,ecasc,exprsc)
                    arcpy.management.CalculateField(outname3paht,ecasc, eca_f_n_exp, "PYTHON3", eca_fix_nulls)
                    print('Calculate ECA Score')
                    arcpy.management.AddField(outname3paht,ecaR,'TEXT')
                    arcpy.management.CalculateField(outname3paht,ecaR,expression,"PYTHON3",codeblk)
                    print('Calculate ECA Rank')
                    arcpy.management.AddField(outname3paht,xcol,'LONG')
                    arcpy.management.CalculateField(outname3paht,xcol,'!Join_Count!',"PYTHON3" )
                    print('Create NO_Crossing Col and populate')
                arcpy.env.workspace = rslt
                
                
                arcpy.management.ValidateJoin(tbl,rev,outname3,rev)
                arcpy.management.JoinField(tbl,rev,outname3,rev,['No_Crossings','SUM_ECA_ha','ECA_Score','ECA_Rank'])
                print('xings and ECA joined to table')
                arcpy.conversion.TableToExcel(tbl,outxls)
                print(wtr +' and ' +tbl+' completed' )

        def build_xlsx (report_folder):
            
            inputlist=[]
            print('BEGIN THE X     L     S    X')

            new_wrkbook=os.path.join(report_folder,(tempname+r'_TEST.xlsx'))
            for f in (os.listdir(report_folder)):
                inputxlsx=os.path.join(report_folder,f)
                # inputxlsx=inputxlsx[1:]
                inputlist.append(inputxlsx)
            # print(inputlist)
            if os.path.exists(report_out):
                # inputlist.remove(report_out)
                print('Template exists in output folder')

            else:
                shutil.copyfile(xlsTemplate,report_out)
                print('xlsx template copied')
            
            # Loop through files in report dir and find watershed spreadsheets
            #create list of input data
            data_lst=[]
            for i in inputlist:
                # print(inputlist)
                if 'Named' in i:
                    # print(i)
                    data_lst.append(i)
                elif 'Tributaries' in i:
                    # print(i)
                    data_lst.append(i)
                elif 'WAU' in i:
                    # print(i)
                    data_lst.append(i) 
            # print(data_lst)

            # #loop trhough input data files and extract data in right order and move it to template xlxs on appropriate tab
            nm_lst=[]
            trib_lst=[]
            wau_lst=[]
            for data in data_lst:
                print(data)
                df=pd.read_excel(data,sheet_name='Compiled_Watershed_Stats_Table_', header=0 )
                # print(wb_name)
                if 'Named' in data:
                    app_lst=nm_lst
                elif 'Tributaries' in data:
                    app_lst=trib_lst
                elif 'WAU' in data:
                    app_lst=wau_lst

                if 'Range_PCNT' not in df.columns:
                    df['Range_PCNT']=''
                    df['Range_PCNT']= df['Range_PCNT'].fillna(0, inplace=True)

                if 'PrivateIR_PCNT' not in df.columns:
                    df['PrivateIR_PCNT']=''
                    df['PrivateIR_PCNT']= df['PrivateIR_PCNT'].fillna(0, inplace=True)

                df.fillna(0, inplace=True)

                for i in range (len(df)):
                    rightorder=[df.loc[i,'OBJECTID'],df.loc[i,'Assess_Uni'],df.loc[i,'RevRepUni'],df.loc[i,'Report_Nam'],df.loc[i,'Report_Typ'],df.loc[i,'RU_Area_ha'],
                          df.loc[i,'RU_Area_km2'],df.loc[i,'RU_Area_m2'],df.loc[i,'MinElev'],df.loc[i,'MaxElev'],df.loc[i,'Elev_Relief'],df.loc[i,'ALPINE_NF_PERCENT'],
                          df.loc[i,'BEC_Score'], df.loc[i,'ECA_Score'],df.loc[i,'DDR_Length_km'],df.loc[i,'DDR_Score'],df.loc[i,'Lake_wetland_adjust_ha'],
                          df.loc[i,'Lake_wetland_Abscence'],df.loc[i,'Terrain_stability_percent'],df.loc[i,'GSC_Score'],df.loc[i,'Percent_steep_coupled'],
                          df.loc[i,'Rds_Extent'],df.loc[i,'RdsStrmB_Ext_KM2'],df.loc[i,'RdsSlps_Ext_KM2'],df.loc[i,'No_Crossings'],df.loc[i,'GOS_Score_Percent'],
                          df.loc[i,'Logged_PCNT'],df.loc[i,'Range_PCNT'],df.loc[i,'PrivateIR_PCNT'],df.loc[i,'Placer_Score'],df.loc[i,'Coal_Lease_PCNT'],
                          df.loc[i,'RUN_DATE']]
                    # rightorder=first+second+third+forth+fifth+sixth+seventh+eigth
                    print(rightorder)
                    app_lst.append(rightorder)
                
                print(nm_lst)
                print(trib_lst)
                print(wau_lst)
                # # print(append_lst)
                

                # Below commented out for testing, un comment once we can confirm that testing is complete 

            #load template
            wb=opxl.load_workbook(filename=report_out)
            sheet_names=wb.sheetnames
            sheet_names.remove('README')
            print(sheet_names)
            
            for sheet_name in sheet_names:
                sheet=wb[sheet_name]
                print('======',sheet_name,'======')
                if sheet_name=='Named Watershed':
                    append_lst=nm_lst
                elif sheet_name=='Tributary Watersheds':
                    append_lst=trib_lst
                elif sheet_name=='Watershed Assessment Units':
                    append_lst=wau_lst
                r=2
                for app in append_lst:
                    print('======Processing row #',str(r),'======')
                    ind=1
                    for a in app:
                        sheet.cell(row=r,column=ind,value=a)
                        ind=ind+1
                    r=r+1      

            wb.save(filename=report_out)
            wb=opxl.load_workbook(filename=report_out,data_only=True)
            wb.save(filename=report_out2)

            print(os.path.isabs(report_out2))

            excel=win32.Dispatch('Excel.Application')
            wb2=excel.Workbooks.Open(report_out2)
            wb2.Save()
            wb2.Close()
            excel.Quit()

            wb2=excel.Workbooks.Open(report_out)
            wb2.Save()
            wb2.Close()
            excel.Quit()
            print(f'{sheet_name} Excel closed')



        def rejoin (inp):
            arcpy.env.workspace =inp
            arcpy.env.overwriteOutput = True
            nmw=arcpy.ListFeatureClasses('*Named*')
            tribw=arcpy.ListFeatureClasses('*Tributaries*')
            wauw=arcpy.ListFeatureClasses('*WAU*')
            print(nmw, tribw, wauw)
            for n in nmw:
                if n.endswith('not_final'):
                    nmw=n
            print(nmw)
            for t in tribw:
                if t.endswith('not_final'):
                    tribw=t
            print(tribw)
            for w in wauw:
                if w.endswith('not_final'):
                    wauw=w
            print(wauw)
            unq='RevRepUni'
            print(today)
            nmsht=os.path.join(report_out,r'T_Named_Watershed$_')
            trbsht=os.path.join(report_out,r'T_Tributary_Watersheds$_')
            wausht=os.path.join(report_out,r'T_Watershed_Assessment_Units$_')

            print(nmsht)
            if arcpy.Exists(nmw):
                unq1=[f.name for f in arcpy.ListFields(nmw,'*RevRepuni')][0]
                print(unq1)
                arcpy.management.ValidateJoin(nmw,unq1,nmsht,unq )
                nm_watershed=arcpy.management.AddJoin(nmw,unq1,nmsht,unq )
                print('Named Joined')
                arcpy.management.CopyFeatures(nm_watershed,('Compiled_Watershed_Features_Named_'+str(today)))
                print('new named')
                arcpy.management.Delete(nmw)

            if arcpy.Exists(tribw):
                unq2=[f.name for f in arcpy.ListFields(tribw,'*RevRepuni')][0]
                print(unq2)
                arcpy.management.ValidateJoin(tribw,unq2,trbsht,unq )
                trib_watershed=arcpy.management.AddJoin(tribw,unq2,trbsht,unq )
                print('Tribs Joined')
                arcpy.management.CopyFeatures(trib_watershed,('Compiled_Watershed_Features_Tributaries_'+str(today)))
                print('new trib')
                arcpy.management.Delete(tribw)

            if arcpy.Exists(wauw):
                unq3=[f.name for f in arcpy.ListFields(wauw,'*RevRepuni')][0]
                print(unq3)
                arcpy.management.ValidateJoin(wauw,unq3,wausht,unq )
                wau_watershed=arcpy.management.AddJoin(wauw,unq3,wausht,unq )
                print('WAU Joined')
                arcpy.management.CopyFeatures(wau_watershed,('Compiled_Watershed_Features_WAU_'+str(today)))
                print('new wau')
                arcpy.management.Delete(wauw)
            
            
            nmdel=arcpy.ListFeatureClasses('*Named_Watershed')
            tribdel=arcpy.ListFeatureClasses('*Features_Tributaries')
            waudel=arcpy.ListFeatureClasses('*Features_WAU')
            aois=arcpy.ListFeatureClasses('*in_AOI')
            print(nmdel)
            print(tribdel)
            print(waudel)
            print(aois)
            arcpy.management.Delete(nmdel[0])
            arcpy.management.Delete(tribdel[0])
            arcpy.management.Delete(waudel[0])
            for a in aois:
                arcpy.management.Delete(a)
            print('REJOIN, FINAL Feature class Created')

        def maps(proProj):
            #get final watersehd data 
            arcpy.env.workspace= outrslt
            print(today)
            namelook='*Named_'+str(year)+'*'
            triblook='*Tributaries_'+str(year)+'*'
            waulook='*WAU_'+str(year)+'*'

            #set project aprx as working aprx, and select results maps
            aprx=arcpy.mp.ArcGISProject(aprxname)
            rslt_map=aprx.listMaps('*WHPOR Results Map*')
            rslt_map=rslt_map[0]
            print('-----------------------name of map-----------------------')
            print(rslt_map.name)
            print('-----------------------name of map-----------------------')

            #get all layers that need data sources changed to new outputs
            named_map_lyrs=rslt_map.listLayers('*Watershed*')
            Trib_map_lyrs=rslt_map.listLayers('*Tributaries*')
            WAU_map_lyrs=rslt_map.listLayers('*WAU*')
            xing_den_map_lyrs=rslt_map.listLayers('Stream Crossing Categories')
            print(len(named_map_lyrs))
            print(len(Trib_map_lyrs))
            print(len(WAU_map_lyrs))
            print(len(xing_den_map_lyrs))

            changelst=[]
            rsltlst=[]
            #get the named watershed final FC and rename all fields to remove any prefixes 
            rslt_name=arcpy.ListFeatureClasses(namelook)
            if len(rslt_name)>0:
                rslt_name=rslt_name[0]
                rsltlst.append(rslt_name)
                changelst.append(named_map_lyrs)
            else:
                for f in named_map_lyrs:
                    rslt_map.removeLayer(f)
                    print('map layer deleted no named watershed')

            rslt_trib=arcpy.ListFeatureClasses(triblook)
            if len(rslt_trib)>0:
                rslt_trib=rslt_trib[0]
                rsltlst.append(rslt_trib)
                changelst.append(Trib_map_lyrs)
            else:
                for f in Trib_map_lyrs:
                    rslt_map.removeLayer(f)
                    print('map layer deleted no tribs')

            rslt_wau=arcpy.ListFeatureClasses(waulook)
            if len(rslt_wau)>0:
                rslt_wau=rslt_wau[0]
                rsltlst.append(rslt_wau)
                changelst.append(WAU_map_lyrs)
            else:
                for f in WAU_map_lyrs:
                    rslt_map.removeLayer(f)
                    print('map layer deleted no WAU')

            # # if aprx does not exist copy the template file
            # if os.path.exists(aprxname):
            #     print(WatershedName+' APRX exists')
            # else:
            #     aprx=arcpy.mp.ArcGISProject(aprxtemp)
            #     aprx.saveACopy(aprxname)
            # aprx.save()
                        
            #standardize field name for resultatnt FCs
            field_names=['ECA_Rank', 'Riparian_Hazard_Score', 'Sediment_Hazard_Score', 'Streamflow_Hazard_Score']
            # og_f_name=[]
            for fc in rsltlst:
                #list fields for ECA_Rank,  Riparian_Hazard_Score, Sediment_Hazard_Score, Streamflow_Hazard_Score
                f1=(arcpy.ListFields(fc, '*ECA_Rank')[0]).name
                f2=(arcpy.ListFields(fc, '*Riparian_Hazard_Score')[0]).name
                f3=(arcpy.ListFields(fc, '*Sediment_Hazard_Score')[0]).name
                f4=(arcpy.ListFields(fc, '*Streamflow_Hazard_Score')[0]).name
                og_f_name=[f1, f2, f3, f4]
            
                for (fn, og) in zip (field_names, og_f_name):
                    arcpy.management.AlterField(in_table=fc, field=og, new_field_name=fn, new_field_alias=fn)
                    print('field altered')
               
            for (lyrs,nm) in zip (changelst,rsltlst):
                print(nm)
                for l in lyrs:
                    origConnPropDict = l.connectionProperties
                    newConnPropDict = {'connection_info': {'database': outrslt},
                            'dataset':nm,
                            'workspace_factory': 'File Geodatabase'}
                    l.updateConnectionProperties(origConnPropDict, newConnPropDict)
                    print(l)
                print(f"Attempting to save .aprx file at: {aprx.filePath}")
                aprx.save()

            # update xing connection properties 
            origConnPropDict = xing_den_map_lyrs[0].connectionProperties
            newConnPropDict = {'connection_info': {'database': outrslt},
                            'dataset':rslt_name,
                            'workspace_factory': 'File Geodatabase'}
            xing_den_map_lyrs[0].updateConnectionProperties(origConnPropDict, newConnPropDict)

            print('WAU watershed connections changed')
            print(f"Attempting to save .aprx file at: {aprx.filePath}")
            aprx.save()

            #set map elemnts
            lyout=aprx.listLayouts('WHPOR Results Map')[0]
            mfrm=lyout.listElements("MAPFRAME_ELEMENT")[0]
            
            title=lyout.listElements('TEXT_ELEMENT','Title')[0]
            title.text=WatershedName+':\nWHPOR Results'
            # scale=int(mfrm.camera.scale)
            scaleBar = lyout.listElements("MAPSURROUND_ELEMENT", 'Alternating Scale Bar')[0]
            print('Title updated')

            #zoom to watershed AOI
            unq1=[f.name for f in arcpy.ListFields(named_map_lyrs[0],'*RevRepuni')][0]
            exprs=unq1+' IS NOT NULL'
            arcpy.management.SelectLayerByAttribute(named_map_lyrs[0],'NEW_SELECTION',exprs)
            # print(arcpy.management.GetCount(named_map_lyrs[0]))
            mfrm.camera.setExtent(mfrm.getLayerExtent(named_map_lyrs[0]))
            scale=int(mfrm.camera.scale)
            arcpy.SelectLayerByAttribute_management(named_map_lyrs[0], "CLEAR_SELECTION")
            print(scale)
            print(len(str(scale)))
            #round map scale 
            if len(str(scale)) == 4:
                new_scale=math.ceil(scale/500)*500
            elif len(str(scale)) == 5:
                new_scale=math.ceil(scale/5000)*5000
            elif len(str(scale)) == 6:
                new_scale=math.ceil(scale/50000)*50000
            elif len(str(scale)) == 7:
                new_scale=math.ceil(scale/500000)*500000
            elif len(str(scale)) == 8:
                new_scale=math.ceil(scale/5000000)*5000000
            else:
                new_scale=scale
            print(new_scale)
            # Apply 10% zoom-out and re-round to a clean map scale.
            zoomed_scale=int(new_scale*1.10)
            if len(str(zoomed_scale)) == 4:
                final_scale=math.ceil(zoomed_scale/500)*500
            elif len(str(zoomed_scale)) == 5:
                final_scale=math.ceil(zoomed_scale/5000)*5000
            elif len(str(zoomed_scale)) == 6:
                final_scale=math.ceil(zoomed_scale/50000)*50000
            elif len(str(zoomed_scale)) == 7:
                final_scale=math.ceil(zoomed_scale/500000)*500000
            elif len(str(zoomed_scale)) == 8:
                final_scale=math.ceil(zoomed_scale/5000000)*5000000
            else:
                final_scale=zoomed_scale

            #set map scale
            mfrm.camera.scale=int(final_scale)
            print('scale rounded from ',scale,' to ',new_scale)
            print('scale after 10% zoom out ',final_scale)
            # Constrain scale bar to the map frame so it cannot run off-frame.
            scaleBar = lyout.listElements("MAPSURROUND_ELEMENT", 'Alternating Scale Bar')[0]
            # Re-apply a known labeled alternating style if available.
            try:
                if hasattr(scaleBar, 'applyStyleItem'):
                    applied_style = False
                    style_names = ['ArcGIS 2D', 'ArcGIS 2D (System)', 'ArcGIS']
                    style_classes = ['Scale_Bar', 'Scale_bar', 'SCALE_BAR', 'Scale Bar']
                    style_wildcards = ['Alternating Scale Bar*', 'Double Alternating Scale Bar*', 'Alternating*', '*Scale Bar*']
                    for st in style_names:
                        if applied_style:
                            break
                        for sc in style_classes:
                            if applied_style:
                                break
                            for wc in style_wildcards:
                                try:
                                    items = aprx.listStyleItems(st, sc, wc)
                                    if items:
                                        scaleBar.applyStyleItem(items[0])
                                        applied_style = True
                                        print('Scale bar style applied:', st, sc, items[0].name)
                                        break
                                except Exception:
                                    pass
                    if not applied_style:
                        print('No labeled scale bar style item found; keeping existing style')
            except Exception as e:
                print('Scale bar style apply skipped:', e)

            frame_left = mfrm.elementPositionX
            frame_bottom = mfrm.elementPositionY
            frame_right = frame_left + mfrm.elementWidth
            frame_top = frame_bottom + mfrm.elementHeight

            left_margin = 0.05
            right_margin = 0.10
            y_margin = 0.05
            target_y = 2.45
            right_label_buffer = 0.70
            left_limit = frame_left + left_margin
            frame_right_limit = frame_right - right_margin
            right_limit = frame_right_limit

            # Align scale bar right edge with north arrow x when available.
            north_arrow_elements = lyout.listElements("MAPSURROUND_ELEMENT", "*North*")
            if not north_arrow_elements:
                north_arrow_elements = lyout.listElements("MAPSURROUND_ELEMENT", "*north*")
            if north_arrow_elements:
                north_arrow_x = north_arrow_elements[0].elementPositionX
                if left_limit + 0.5 < north_arrow_x <= frame_right_limit:
                    right_limit = north_arrow_x
                    print('Scale bar right edge aligned to north arrow x:', round(north_arrow_x, 3))
                else:
                    print('North arrow x outside safe range; using frame right limit')
            else:
                print('North arrow element not found; using frame right limit')

            if right_limit <= left_limit + 0.5:
                right_limit = frame_right_limit

            label_safe_right = right_limit - right_label_buffer
            if label_safe_right <= left_limit + 0.5:
                label_safe_right = frame_right_limit - right_label_buffer

            max_allowed_width = max(0.8, label_safe_right - left_limit)
            target_width = max(0.8, mfrm.elementWidth * 0.35)
            desired_width = min(target_width, max_allowed_width)
            target_total_km = max(0.2, (final_scale * desired_width * 0.0254) / 1000.0)
            division_km = max(0.05, target_total_km / 4.0)

            # Keep a visible alternating pattern and prevent width auto-growth.
            try:
                scalebar_cim = scaleBar.getDefinition('V3')
                if hasattr(scalebar_cim, 'fittingStrategy'):
                    try:
                        scalebar_cim.fittingStrategy = 'AdjustDivision'
                    except Exception:
                        scalebar_cim.fittingStrategy = 0
                if hasattr(scalebar_cim, 'divisions'):
                    scalebar_cim.divisions = 4
                if hasattr(scalebar_cim, 'division'):
                    scalebar_cim.division = division_km
                if hasattr(scalebar_cim, 'subdivisions'):
                    scalebar_cim.subdivisions = 0
                if hasattr(scalebar_cim, 'labelFrequency'):
                    try:
                        scalebar_cim.labelFrequency = 'Divisions'
                    except Exception:
                        scalebar_cim.labelFrequency = 3
                if hasattr(scalebar_cim, 'displayFirstOutside'):
                    scalebar_cim.displayFirstOutside = True
                if hasattr(scalebar_cim, 'displayLastOutside'):
                    scalebar_cim.displayLastOutside = True
                if hasattr(scalebar_cim, 'markFrequency'):
                    try:
                        scalebar_cim.markFrequency = 'None'
                    except Exception:
                        scalebar_cim.markFrequency = 0
                if hasattr(scalebar_cim, 'divisionMarkHeight'):
                    scalebar_cim.divisionMarkHeight = 0
                if hasattr(scalebar_cim, 'subdivisionMarkHeight'):
                    scalebar_cim.subdivisionMarkHeight = 0
                if hasattr(scalebar_cim, 'labelGap'):
                    scalebar_cim.labelGap = 1.5
                if hasattr(scalebar_cim, 'unitLabelGap'):
                    scalebar_cim.unitLabelGap = 1.5
                if hasattr(scalebar_cim, 'units'):
                    try:
                        scalebar_cim.units = {'uwkid': 9036}
                    except Exception:
                        try:
                            scalebar_cim.units = {'wkid': 9036}
                        except Exception:
                            pass
                if hasattr(scalebar_cim, 'showLabels'):
                    scalebar_cim.showLabels = True
                if hasattr(scalebar_cim, 'showDivisionLabels'):
                    scalebar_cim.showDivisionLabels = True
                if hasattr(scalebar_cim, 'showFirstLabel'):
                    scalebar_cim.showFirstLabel = True
                if hasattr(scalebar_cim, 'showLastLabel'):
                    scalebar_cim.showLastLabel = True
                if hasattr(scalebar_cim, 'showUnitLabel'):
                    scalebar_cim.showUnitLabel = True
                if hasattr(scalebar_cim, 'unitLabel'):
                    scalebar_cim.unitLabel = 'km'
                if hasattr(scalebar_cim, 'unitLabelPosition'):
                    try:
                        scalebar_cim.unitLabelPosition = 'AfterBar'
                    except Exception:
                        scalebar_cim.unitLabelPosition = 4
                if hasattr(scalebar_cim, 'labelSymbol') and hasattr(scalebar_cim, 'unitLabelSymbol'):
                    if scalebar_cim.unitLabelSymbol:
                        scalebar_cim.labelSymbol = scalebar_cim.unitLabelSymbol
                if hasattr(scalebar_cim, 'numberFormat') and scalebar_cim.numberFormat:
                    if hasattr(scalebar_cim.numberFormat, 'roundingValue'):
                        if division_km >= 1:
                            scalebar_cim.numberFormat.roundingValue = 1
                        elif division_km >= 0.1:
                            scalebar_cim.numberFormat.roundingValue = 0.1
                        else:
                            scalebar_cim.numberFormat.roundingValue = 0.01
                if hasattr(scalebar_cim, 'labelPosition'):
                    try:
                        scalebar_cim.labelPosition = 'Above'
                    except Exception:
                        scalebar_cim.labelPosition = 0
                available_label_props = [
                    p for p in [
                        'displayFirstOutside', 'displayLastOutside', 'labelFrequency', 'labelPosition',
                        'labelSymbol', 'unitLabel', 'unitLabelPosition', 'unitLabelSymbol', 'numberFormat'
                    ] if hasattr(scalebar_cim, p)
                ]
                print('Scale bar CIM label props available:', available_label_props)
                scaleBar.setDefinition(scalebar_cim)
                print('Scale bar divisions, fitting strategy, and labels set (division_km=', round(division_km, 3), ')')
            except Exception as e:
                print('Scale bar division update skipped:', e)

            try:
                scaleBar.setAnchor("BOTTOM_RIGHT_CORNER")
            except Exception:
                pass

            scaleBar.elementWidth = desired_width
            scaleBar.elementPositionX = right_limit
            min_y = frame_bottom + y_margin
            max_y = max(min_y, frame_top - scaleBar.elementHeight - y_margin)
            scaleBar.elementPositionY = max(min_y, min(target_y, max_y))

            # Re-clamp after CIM updates in case the style recalculates width.
            if scaleBar.elementWidth > max_allowed_width:
                scaleBar.elementWidth = max_allowed_width

            left_edge = scaleBar.elementPositionX - scaleBar.elementWidth
            if left_edge < left_limit:
                scaleBar.elementWidth = max(0.8, scaleBar.elementPositionX - left_limit)
                left_edge = scaleBar.elementPositionX - scaleBar.elementWidth

            print('Scale bar constrained to frame: left', round(left_edge, 3), 'right', round(scaleBar.elementPositionX, 3), 'target_right', round(right_limit, 3), 'frame_right', round(frame_right - right_margin, 3), 'width', round(scaleBar.elementWidth, 3))
            print(f"Attempting to save .aprx file at: {aprx.filePath}")
            aprx.save()  
            print('Scale Bar Adjusted')
            print('export map')
            lyout.exportToPDF(mapout)
            print('Layout exported to ', mapout)
            print(f"Attempting to save .aprx file at: {aprx.filePath}")
            aprx.save()  

        def copDevs(final_location):
            if not os.path.exists(final_location):
                os.makedirs(final_location)

            map_nm_like=(WatershedName+r' WHPOR Results Map '+str(year))
            for root, dirs,files in os.walk(os.path.join(BaseFolder,r'3_Maps')):
                for file in files:
                    if file.startswith(map_nm_like):
                        shutil.copy(mapout,os.path.join(final_location,mapname))
            
            rprt_nm_like=(watershedname+r'_Compiled_Watershed_Hazard_Summaries_'+str(year))
            for root, dirs,files in os.walk(rprtFolder):
                for file in files:
                    if file.startswith(rprt_nm_like):
                        shutil.copy(os.path.join(rprtFolder,file),os.path.join(final_location,tempname))

            print('MAP and Spreadsheet copied to final location')
            print('\n\n\n\n\n\n\n\n\n\n\n')
            print('==========================================================================================================================================================================================')
            print('****************FINAL DELIVERABLES****************FINAL DELIVERABLES****************FINAL DELIVERABLES****************FINAL DELIVERABLES****************FINAL DELIVERABLES****************')
            print('==========================================================================================================================================================================================')
            print(mapout)
            print(report_out)
            print(outrslt)
            print('Client Directory with deliverables in place')
            print(clientdir)
            print('==========================================================================================================================================================================================')
            print('****************FINAL DELIVERABLES****************FINAL DELIVERABLES****************FINAL DELIVERABLES****************FINAL DELIVERABLES****************FINAL DELIVERABLES****************')
            print('==========================================================================================================================================================================================')
        #=====Call Functions=====
        move_CEA(rsltGdb)
        xing_eca (outrslt,inputgdb)
        build_xlsx(rprtFolder) 
        rejoin(outrslt)
        maps(aprxtemp)
        copDevs(clientdir)

